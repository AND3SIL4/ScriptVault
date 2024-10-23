import pandas as pd  # type: ignore
import numpy as np  # type: ignore
from typing import Optional
import os
from openpyxl import load_workbook  # type: ignore


class Consecutivo:
    """Clase para manejar la información de coaseguros"""

    def __init__(
        self,
        file_path: str,
        sheet_name: str,
        inconsistencies_file: str,
        exception_file: str,
        consecutivo_sap_file: str,
    ):
        self.path_file = file_path
        self.sheet_name = sheet_name
        self.inconsistencies_file = inconsistencies_file
        self.exception_file = exception_file
        self.consecutivo_sap_file = consecutivo_sap_file

    def read_excel(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """Method for returning a data frame"""
        return pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

    def save_inconsistencies_file(self, df: pd.DataFrame, new_sheet: str) -> bool:
        if os.path.exists(self.inconsistencies_file):
            with pd.ExcelFile(self.inconsistencies_file, engine="openpyxl") as xls:
                if new_sheet in xls.sheet_names:
                    existing = pd.read_excel(
                        xls, engine="openpyxl", sheet_name=new_sheet
                    )
                    df = pd.concat([existing, df], ignore_index=True)

            with pd.ExcelWriter(
                self.inconsistencies_file,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as writer:
                df.to_excel(writer, sheet_name=new_sheet, index=False)
                return True
        else:
            return False

    def excel_col_name(self, number) -> str:
        """Method to convert (1-based) to Excel column name"""
        result = ""
        while number > 0:
            number, reminder = divmod(number - 1, 26)
            result = chr(65 + reminder) + result
        return result

    def validate_inconsistencies(
        self, df: pd.DataFrame, col_idx, sheet_name: str
    ) -> str:
        """Method to validate the inconsistencies before append in a inconsistencies file"""
        if not df.empty:
            df = df.copy()
            if isinstance(col_idx, int):
                df[f"COORDENADAS"] = df.apply(
                    lambda row: f"{self.excel_col_name(col_idx + 1)}{row.name + 2}",
                    axis=1,
                )
            else:
                for i in col_idx:
                    df[f"COORDENADAS_{i + 2}"] = df.apply(
                        lambda row: f"{self.excel_col_name(i + 1)}{row.name + 2}",
                        axis=1,
                    )
            self.save_inconsistencies_file(df, sheet_name)
            return "SUCCESS: Inconsistencies guardadas correctamente"
        else:
            return "INFO: Validacion realizada, no se encontraron inconsistencias"

    def filter_file(
        self, data_frame: pd.DataFrame, cut_off_date: str, col_idx: int
    ) -> pd.DataFrame:
        """Method to filter the file according to the defined criteria"""
        ## Parse date and get the
        date: pd.Timestamp = pd.to_datetime(
            cut_off_date, format="%d/%m/%Y", errors="coerce"
        )
        year = date.year
        month = date.month
        ## Filter file
        filtered_df: pd.DataFrame = data_frame[
            (data_frame.iloc[:, col_idx].dt.month == month)
            & (data_frame.iloc[:, col_idx].dt.year == year)
        ]
        return filtered_df  ##Return df filtered

    def consecutivo(self, cut_off_date: str) -> str:
        ## Initial data frames
        pagos_file: pd.DataFrame = self.read_excel(self.path_file, self.sheet_name)
        consecutivo_file: pd.DataFrame = self.read_excel(
            self.consecutivo_sap_file, "NUEMRO DE PAGO"
        )
        list_df: pd.DataFrame = self.read_excel(self.exception_file, "CONSECUTIVO SAP")
        consecutivo_df = self.filter_file(consecutivo_file, cut_off_date, 0)
        pagos_df = self.filter_file(pagos_file, cut_off_date, 72)

        ## Local variables
        initial_consecutivo: int = int(list_df.iloc[0, 1])
        final_consecutivo: int = int(list_df.iloc[0, 2])
        pending_list: list[int] = list_df.iloc[:, 0].dropna().astype(int).to_list()
        consecutivos_pagos: list = (
            pagos_df.iloc[:, 73].drop_duplicates().to_list()
        )  ##! From Pagos file
        length_consecutivo_pagos: int = int(len(consecutivos_pagos))

        consecutivos_to_validate: list[int] = []
        for consecutivo in range(length_consecutivo_pagos):
            final_consecutivo += 1
            consecutivos_to_validate.append(final_consecutivo)

        ## Validate the if the pending list is in the current consecutivo list
        missing_before_values = [
            value for value in pending_list if value not in consecutivos_to_validate
        ]
        ## Create new data frames to cross over files
        consecutivo_cross: pd.DataFrame = pd.DataFrame(
            consecutivos_to_validate, columns=["CONSECUTIVO_FROM_CONSECUTIVO"]
        )
        consecutivo_pagos_cross: pd.DataFrame = pd.DataFrame(
            consecutivos_pagos, columns=["CONSECUTIVO_FROM_PAGOS"]
        )

        ## Cross files
        consecutivo_from_consecutivo_merged: pd.DataFrame = consecutivo_cross.merge(
            consecutivo_df,
            left_on=consecutivo_cross.columns[0],
            right_on=consecutivo_df.columns[1],
            how="left",
        )
        consecutivos_from_pagos_merged: pd.DataFrame = consecutivo_pagos_cross.merge(
            consecutivo_df,
            left_on=consecutivo_pagos_cross.columns[0],
            right_on=consecutivo_df.columns[1],
            how="left",
        )
        ## Add a column to indicate if the consecutivo from consecutivo is in the consecutivo from pagos
        consecutivo_from_consecutivo_merged["is_valid"] = (
            consecutivo_from_consecutivo_merged.iloc[:, 0].isin(
                consecutivos_from_pagos_merged.iloc[:, 0]
            )
        )
        ## Filter the cases that does not match
        inconsistencies_validation: pd.DataFrame = consecutivo_from_consecutivo_merged[
            ~consecutivo_from_consecutivo_merged["is_valid"]
        ]
        ## Don't take in count "RED ASISTENCIAL" in inconsistencies
        inconsistencies_validation = inconsistencies_validation[
            inconsistencies_validation.iloc[:, 3] != "RED ASISTENCIAL"
        ]
        append_list = (
            missing_before_values
            + inconsistencies_validation.iloc[:, 0].dropna().to_list()
        )
        data_updated: bool = self.update_data(
            consecutivos_to_validate[0], consecutivos_to_validate[-1], append_list
        )
        ## Save the inconsistencies
        return (
            self.validate_inconsistencies(
                inconsistencies_validation, 73, "ValidacionConsecutivo"
            ),
            data_updated,
        )

    def update_data(
        self, consecutivo_inicial: int, consecutivo_final: int, lista_consecutivos: list
    ) -> None:
        # Cargar el archivo existente
        book = load_workbook(self.exception_file)
        sheet = book["CONSECUTIVO SAP"]

        # Limpiar los datos antiguos en la columna 1 (consecutivos pendientes)
        max_row = sheet.max_row  # Obtiene el número máximo de filas usadas
        for row in range(2, max_row + 1):  # Empieza en la fila 2
            sheet.cell(row=row, column=1).value = None  # Borra los valores antiguos

        # Actualizar valores
        sheet.cell(row=2, column=2).value = consecutivo_inicial  # Consecutivo inicial
        sheet.cell(row=2, column=3).value = consecutivo_final  # Consecutivo final

        # Actualizar lista de consecutivos pendientes
        row = 1
        for consecutivo in lista_consecutivos:
            row += 1
            sheet.cell(row=row, column=1).value = consecutivo

        # Guardar cambios
        book.save(self.exception_file)
        return True


##* INITIALIZE THE VARIABLE TO INSTANCE THE MAIN CLASS
consecutivo: Optional[Consecutivo] = None


##* CALL THE MAIN FUNCTION WITH THE MAIN PARAMS
def main(params: dict) -> bool:
    try:
        global consecutivo

        ## Get the variables
        file_path: str = params.get("file_path")
        sheet_name: str = params.get("sheet_name")
        inconsistencies_file: str = params.get("inconsistencies_file")
        exception_file: str = params.get("exception_file")
        consecutivo_sap_file: str = params.get("consecutivo_sap_file")

        ## Pass the values to the constructor in the main class
        consecutivo = Consecutivo(
            file_path,
            sheet_name,
            inconsistencies_file,
            exception_file,
            consecutivo_sap_file,
        )
        return True
    except Exception as e:
        return f"ERROR: {e}"


def validate_consecutivo_sap(params: dict) -> str:
    try:
        ## Set local variables
        cut_off_date: str = params.get("cut_off_date")
        validation: str = consecutivo.consecutivo(cut_off_date)
        return validation
    except Exception as e:
        return f"ERROR: {e}"


if __name__ == "__main__":
    params = {
        "file_path": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\TempFolder\BASE DE PAGOS.xlsx",
        "sheet_name": "PAGOS",
        "inconsistencies_file": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\OutputFolder\Inconsistencias\InconBasePagos.xlsx",
        "exception_file": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\InputFolder\EXCEPCIONES BASE PAGOS.xlsx",
        "consecutivo_sap_file": r"C:\ProgramData\AutomationAnywhere\Bots\Logs\AD_RCSN_SabanaPagosYBasesParaSinestralidad\InputFolder\CONSECUTIVO SAP 2023.xlsx",
    }
    main(params)
    params = {"cut_off_date": "30/07/2024"}
    print(validate_consecutivo_sap(params))
